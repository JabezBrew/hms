import logging

from celery import shared_task
from django.db import transaction
from django.db.models import Sum, Q
from django.utils import timezone
from decimal import Decimal

from apps.billing.models import (
    Claim,
    Invoice,
    InvoiceItem,
    Service,
    ServiceCategory,
    PayerServiceCode,
    PayerServiceCodeImportJob,
    PaymentIntent,
    PSPWebhookEvent,
    Payment,
    Receipt,
    SettlementBatch,
    SettlementLine,
    NHISClaimBatch,
    NHISClaimExportJob,
    ClaimValidationIssue,
    RemittanceImportJob,
    RemittanceLine,
    InsurancePosting,
)
from apps.fhir_client.client import fhir_client
from apps.fhir_client.utils import generate_fhir_id, create_reference
from apps.interop.crypto import decrypt_payload, encrypt_payload
from apps.billing.psp import get_psp_adapter

logger = logging.getLogger(__name__)


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def create_fhir_claim_for_claim(self, claim_id: str) -> None:
    """
    Async FHIR claim creation.

    SECURITY/PERF:
    - Never block request threads on external I/O.
    - Do not log PHI; log claim IDs only.
    """
    claim = Claim.objects.select_related(
        'invoice__patient',
        'invoice__patient_insurance__plan__provider',
    ).prefetch_related(
        'invoice__items__service',
    ).filter(id=claim_id).first()
    if not claim:
        logger.warning("FHIR claim task: claim not found (id=%s)", claim_id)
        return

    invoice = claim.invoice
    patient = getattr(invoice, 'patient', None)
    patient_fhir_id = getattr(patient, 'fhir_patient_id', None) if patient else None
    if not patient_fhir_id:
        logger.info("FHIR claim task: patient has no fhir_patient_id (claim_id=%s)", claim_id)
        return

    # Idempotency: if we already have a fhir_claim_id, don't create another.
    if claim.fhir_claim_id:
        return

    insurance_display = None
    if invoice.patient_insurance and invoice.patient_insurance.plan and invoice.patient_insurance.plan.provider:
        insurance_display = f"{invoice.patient_insurance.plan.provider.name} - {invoice.patient_insurance.plan.name}"

    claim_data = {
        "resourceType": "Claim",
        "id": generate_fhir_id(),
        "status": "active",
        "type": {
            "coding": [
                {
                    "system": "http://terminology.hl7.org/CodeSystem/claim-type",
                    "code": "professional",
                    "display": "Professional",
                }
            ]
        },
        "use": "claim",
        "patient": create_reference("Patient", patient_fhir_id),
        "created": invoice.invoice_date.isoformat() if invoice.invoice_date else None,
        "provider": {"display": "Hospital Management System"},
        "priority": {
            "coding": [
                {
                    "system": "http://terminology.hl7.org/CodeSystem/processpriority",
                    "code": "normal",
                }
            ]
        },
        "insurance": [
            {
                "sequence": 1,
                "focal": True,
                "coverage": {"display": insurance_display} if insurance_display else {"display": "Insurance"},
            }
        ],
        "item": [],
    }

    for idx, item in enumerate(invoice.items.all()):
        service = getattr(item, 'service', None)
        claim_data["item"].append(
            {
                "sequence": idx + 1,
                "productOrService": {
                    "coding": [
                        {
                            "system": "http://hospital.example.org/fhir/service",
                            "code": getattr(service, 'code', None) or "unknown",
                            "display": getattr(service, 'name', None) or "Service",
                        }
                    ]
                },
                "unitPrice": {
                    "value": float(item.unit_price),
                    "currency": "GHS",
                },
                "net": {
                    "value": float(item.total_price),
                    "currency": "GHS",
                },
            }
        )

    fhir_claim = fhir_client.create_resource("Claim", claim_data)
    fhir_claim_id = fhir_claim.get("id")
    if not fhir_claim_id:
        raise RuntimeError("FHIR claim creation returned no id")

    with transaction.atomic():
        # Re-read with row lock to avoid concurrent writes from multiple tasks.
        locked = Claim.objects.select_for_update().select_related('invoice').get(id=claim.id)
        if locked.fhir_claim_id:
            return
        locked.fhir_claim_id = fhir_claim_id
        locked.save(update_fields=['fhir_claim_id', 'updated_at'])

        if locked.invoice and not locked.invoice.fhir_claim_id:
            locked.invoice.fhir_claim_id = fhir_claim_id
            locked.invoice.save(update_fields=['fhir_claim_id', 'updated_at'])


def _recompute_and_persist_invoice_status(invoice) -> None:
    """
    Mirror billing.views CFO-grade status derivation without importing DRF views.
    """
    totals = Payment.objects.filter(invoice=invoice, status='posted').aggregate(
        patient_paid=Sum('amount', filter=Q(payer='patient')),
        insurance_paid=Sum('amount', filter=Q(payer='insurance')),
    )
    patient_paid = totals.get('patient_paid') or Decimal('0.00')
    insurance_paid = totals.get('insurance_paid') or Decimal('0.00')

    patient_responsibility = Decimal(str(getattr(invoice, 'patient_responsibility', 0) or 0))
    insurance_amount = Decimal(str(getattr(invoice, 'insurance_amount', 0) or 0))
    patient_balance_due = patient_responsibility - Decimal(str(patient_paid))
    insurance_balance_due = insurance_amount - Decimal(str(insurance_paid))
    total_balance_due = patient_balance_due + insurance_balance_due

    next_status = invoice.status
    if invoice.status == 'cancelled':
        next_status = 'cancelled'
    elif total_balance_due <= 0:
        next_status = 'paid'
    else:
        total_paid = Decimal(str(patient_paid)) + Decimal(str(insurance_paid))
        if total_paid > 0:
            next_status = 'partially_paid'
        else:
            due_date = getattr(invoice, 'due_date', None)
            if due_date and patient_balance_due > 0 and due_date < timezone.now().date():
                next_status = 'overdue'
            else:
                next_status = 'pending'

    if invoice.status != next_status:
        invoice.status = next_status
        invoice.save(update_fields=['status', 'updated_at'])


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 5})
def process_psp_webhook_event(self, webhook_event_id: str) -> None:
    """
    Process a stored PSP webhook event and post patient payment idempotently.

    SECURITY:
    - Never log payload contents (may contain phone numbers).
    """
    event = PSPWebhookEvent.objects.filter(id=webhook_event_id).first()
    if not event:
        return

    # Fast path: already processed.
    if event.processing_status != 'pending':
        return

    adapter = get_psp_adapter(event.provider)
    payload_bytes = decrypt_payload(event.payload_encrypted) if event.payload_encrypted else b""
    parsed = adapter.parse_webhook(body_bytes=payload_bytes, headers=event.headers or {})

    now = timezone.now()

    with transaction.atomic():
        locked_event = PSPWebhookEvent.objects.select_for_update().get(id=event.id)
        if locked_event.processing_status != 'pending':
            return

        provider_reference = parsed.provider_reference or locked_event.provider_reference
        client_reference = parsed.client_reference or locked_event.client_reference

        locked_event.provider_reference = provider_reference
        locked_event.client_reference = client_reference

        # Postgres cannot apply FOR UPDATE to the nullable side of an outer join.
        # Lock only the PaymentIntent row; load nullable relations lazily.
        intent_qs = PaymentIntent.objects.select_for_update(of=('self',)).select_related(
            'invoice', 'facility'
        )
        intent = None
        if provider_reference:
            intent = intent_qs.filter(provider=locked_event.provider, provider_reference=provider_reference).first()
        if not intent and client_reference:
            intent = intent_qs.filter(provider=locked_event.provider, client_reference=client_reference).first()

        if not intent:
            locked_event.processing_status = 'ignored'
            locked_event.processed_at = now
            locked_event.save(update_fields=[
                'provider_reference', 'client_reference',
                'processing_status', 'processed_at', 'updated_at'
            ])
            return

        # Update intent status and post payment when succeeded.
        next_status = parsed.status
        if next_status not in dict(PaymentIntent.STATUS_CHOICES):
            next_status = 'pending'

        if next_status == 'succeeded':
            if not intent.payment_id:
                paid_amount = parsed.paid_amount or intent.amount
                paid_at = parsed.paid_at or now

                payment_date = getattr(paid_at, 'date', None)() if callable(getattr(paid_at, 'date', None)) else now.date()

                actor = (
                    getattr(intent, 'initiated_by', None) or
                    getattr(intent.invoice, 'updated_by', None) or
                    getattr(intent.invoice, 'created_by', None)
                )

                payment = Payment(
                    invoice=intent.invoice,
                    payment_date=payment_date,
                    amount=paid_amount,
                    payer='patient',
                    status='posted',
                    payment_method=intent.payment_method,
                    reference_number=provider_reference,
                    notes=f"{intent.provider} PSP payment",
                    cash_session=intent.cash_session,
                    created_by=actor,
                    updated_by=actor,
                )
                payment.full_clean()
                payment.save()

                receipt_number = f"RCP-{now.strftime('%Y%m%d')}-{str(payment.id).split('-')[0].upper()}"
                Receipt.objects.create(
                    receipt_number=receipt_number,
                    payment=payment,
                    created_by=actor,
                    updated_by=actor,
                )

                intent.payment = payment

            intent.status = 'succeeded'
            intent.paid_amount = parsed.paid_amount or intent.paid_amount or intent.amount
            intent.fee_amount = parsed.fee_amount or intent.fee_amount
            intent.paid_at = parsed.paid_at or intent.paid_at or now
            intent.save(update_fields=[
                'status', 'payment', 'paid_amount', 'fee_amount', 'paid_at', 'updated_at'
            ])

            _recompute_and_persist_invoice_status(intent.invoice)

        elif next_status in {'failed', 'cancelled', 'expired'}:
            intent.status = next_status
            intent.save(update_fields=['status', 'updated_at'])
        else:
            # pending / unknown
            if intent.status not in {'succeeded'}:
                intent.status = 'pending'
                intent.save(update_fields=['status', 'updated_at'])

        locked_event.processing_status = 'processed'
        locked_event.processed_at = now
        locked_event.save(update_fields=[
            'provider_reference', 'client_reference',
            'processing_status', 'processed_at', 'updated_at'
        ])


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def process_settlement_batch(self, settlement_batch_id: str) -> None:
    """
    Parse an uploaded settlement CSV and attempt to match to intents/payments.

    The CSV format varies by provider and exporter. For v1 we support a minimal
    column set:
    - provider_reference (or paylink_id)
    - client_reference (optional)
    - amount_gross (or amount)
    - fee_amount (or fee)
    - amount_net (or net_amount)
    - paid_at (optional)
    """
    import csv
    import io

    batch = SettlementBatch.objects.select_related('facility').filter(id=settlement_batch_id).first()
    if not batch:
        return

    if batch.status != 'pending':
        return

    now = timezone.now()

    with transaction.atomic():
        locked = SettlementBatch.objects.select_for_update().get(id=batch.id)
        if locked.status != 'pending':
            return
        locked.status = 'running'
        locked.save(update_fields=['status', 'updated_at'])

    try:
        raw = decrypt_payload(batch.payload_encrypted) if batch.payload_encrypted else b""
        text = raw.decode('utf-8', errors='replace')
        reader = csv.DictReader(io.StringIO(text))

        lines = []
        for row in reader:
            norm = {str(k or '').strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in (row or {}).items()}
            provider_reference = norm.get('provider_reference') or norm.get('paylink_id') or norm.get('paylinkid') or norm.get('reference')
            client_reference = norm.get('client_reference') or norm.get('clientreference')

            amount_gross = norm.get('amount_gross') or norm.get('amount') or norm.get('gross_amount')
            fee_amount = norm.get('fee_amount') or norm.get('fee') or norm.get('fees') or norm.get('charges')
            amount_net = norm.get('amount_net') or norm.get('net_amount') or norm.get('net')

            def _d(val):
                try:
                    return Decimal(str(val)) if val not in (None, '') else None
                except Exception:
                    return None

            line = SettlementLine(
                batch=batch,
                provider_reference=str(provider_reference) if provider_reference else None,
                client_reference=str(client_reference) if client_reference else None,
                status=str(norm.get('status') or '').strip(),
                amount_gross=_d(amount_gross),
                fee_amount=_d(fee_amount),
                amount_net=_d(amount_net),
                match_status='unmatched',
            )
            lines.append(line)

        SettlementLine.objects.bulk_create(lines, batch_size=500)

        # Match lines
        qs = SettlementLine.objects.filter(batch=batch).select_related('matched_intent', 'matched_payment')
        for line in qs.iterator(chunk_size=500):
            intent = None
            payment = None
            if line.provider_reference:
                intent = PaymentIntent.objects.filter(
                    facility=batch.facility,
                    provider=batch.provider,
                    provider_reference=line.provider_reference,
                ).select_related('payment').first()
                payment = Payment.objects.filter(
                    invoice__facility=batch.facility,
                    reference_number=line.provider_reference,
                    status='posted',
                ).first()
            if not intent and line.client_reference:
                intent = PaymentIntent.objects.filter(
                    facility=batch.facility,
                    provider=batch.provider,
                    client_reference=line.client_reference,
                ).select_related('payment').first()

            match_status = 'unmatched'
            mismatch_reason = ''
            if intent or payment:
                match_status = 'matched'
            if intent and intent.payment_id and payment and payment.id != intent.payment_id:
                match_status = 'mismatch'
                mismatch_reason = 'provider_reference matches multiple payments'

            matched_payment = payment or getattr(intent, 'payment', None)
            SettlementLine.objects.filter(id=line.id).update(
                matched_intent_id=getattr(intent, 'id', None),
                matched_payment_id=getattr(matched_payment, 'id', None),
                match_status=match_status,
                mismatch_reason=mismatch_reason,
            )

            # Best-effort fees reconciliation on intent.
            if intent and line.fee_amount is not None and (intent.fee_amount is None or intent.fee_amount == 0):
                PaymentIntent.objects.filter(id=intent.id).update(fee_amount=line.fee_amount, updated_at=now)

        SettlementBatch.objects.filter(id=batch.id).update(
            status='ready',
            processed_at=now,
            updated_at=now,
        )
    except Exception as e:
        SettlementBatch.objects.filter(id=batch.id).update(
            status='failed',
            error_message=str(e)[:500],
            processed_at=now,
            updated_at=now,
        )
        raise


def _parse_date(val):
    """
    Parse a date from common Ghana-friendly formats.

    Accepted:
    - YYYY-MM-DD (preferred)
    - YYYY/MM/DD
    - DD/MM/YYYY
    - MM/DD/YYYY
    """
    from datetime import datetime

    if val in (None, ''):
        return None
    if hasattr(val, 'date') and not isinstance(val, str):
        # openpyxl may return datetime/date objects.
        try:
            return val.date()
        except Exception:
            pass
    s = str(val).strip()
    if not s:
        return None
    try:
        # date.fromisoformat handles YYYY-MM-DD
        return datetime.fromisoformat(s).date()
    except Exception:
        pass
    for fmt in ("%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    raise ValueError(f"Invalid date: {s}")


def _parse_bool(val, default=True):
    if val in (None, ''):
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ('1', 'true', 't', 'yes', 'y'):
        return True
    if s in ('0', 'false', 'f', 'no', 'n'):
        return False
    return default


def _normalize_header(s):
    return str(s or '').strip().lower().replace(' ', '_').replace('-', '_')


def _read_mapping_rows(*, raw: bytes, file_name: str):
    """
    Return a list of dict rows from CSV or XLSX.

    Expected headers (case-insensitive):
    - service_code (required)
    - external_code (required)
    - effective_from (required)
    - effective_until (optional)
    - is_active (optional)

    Optional (used only when seed_services=True):
    - service_name
    - category_name
    - base_price
    """
    import csv
    import io

    if (file_name or '').lower().endswith(('.xlsx', '.xls')):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None) or []
        keys = [_normalize_header(h) for h in headers]
        rows = []
        for r in rows_iter:
            if not r or all(v in (None, '') for v in r):
                continue
            row = {keys[i]: r[i] if i < len(r) else None for i in range(len(keys))}
            rows.append(row)
        return rows

    # CSV (default)
    text = raw.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        norm = {_normalize_header(k): (v.strip() if isinstance(v, str) else v) for k, v in (row or {}).items()}
        if not norm:
            continue
        # Skip empty lines
        if all(v in (None, '') for v in norm.values()):
            continue
        rows.append(norm)
    return rows


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def process_payer_service_code_import_job(self, import_job_id: str) -> None:
    """
    Async preview parse/validate for payer service code imports.

    SECURITY:
    - No PHI expected, but do not log file contents.
    """
    import json

    job = PayerServiceCodeImportJob.objects.select_related('facility', 'payer').filter(id=import_job_id).first()
    if not job:
        return
    if job.status not in ('pending',):
        return

    now = timezone.now()
    with transaction.atomic():
        locked = PayerServiceCodeImportJob.objects.select_for_update().get(id=job.id)
        if locked.status != 'pending':
            return
        locked.status = 'running'
        locked.save(update_fields=['status', 'updated_at'])

    issues = []
    summary = {
        'rows_total': 0,
        'rows_valid': 0,
        'errors': 0,
        'warnings': 0,
        'would_create_mappings': 0,
        'would_update_mappings': 0,
        'would_create_services': 0,
        'would_create_categories': 0,
        'duplicates_in_file': 0,
    }

    try:
        raw = decrypt_payload(job.payload_encrypted) if job.payload_encrypted else b""
        rows = _read_mapping_rows(raw=raw, file_name=job.file_name or '')
        summary['rows_total'] = len(rows)

        # First pass normalize/validate required fields and collect service codes.
        normalized = []
        service_codes = set()
        seen_keys = set()
        for idx, row in enumerate(rows, start=2):  # csv header is row 1; treat first data as row 2
            service_code = (row.get('service_code') or '').strip() if isinstance(row.get('service_code'), str) else str(row.get('service_code') or '').strip()
            external_code = (row.get('external_code') or '').strip() if isinstance(row.get('external_code'), str) else str(row.get('external_code') or '').strip()
            effective_from_raw = row.get('effective_from')
            effective_until_raw = row.get('effective_until')

            row_errors = []
            if not service_code:
                row_errors.append({'field': 'service_code', 'code': 'required', 'message': 'service_code is required'})
            if not external_code:
                row_errors.append({'field': 'external_code', 'code': 'required', 'message': 'external_code is required'})

            eff_from = None
            eff_until = None
            try:
                eff_from = _parse_date(effective_from_raw)
            except Exception:
                row_errors.append({'field': 'effective_from', 'code': 'invalid', 'message': 'effective_from must be a date'})
            if not eff_from:
                row_errors.append({'field': 'effective_from', 'code': 'required', 'message': 'effective_from is required'})

            if effective_until_raw not in (None, ''):
                try:
                    eff_until = _parse_date(effective_until_raw)
                except Exception:
                    row_errors.append({'field': 'effective_until', 'code': 'invalid', 'message': 'effective_until must be a date'})

            if eff_from and eff_until and eff_until < eff_from:
                row_errors.append({'field': 'effective_until', 'code': 'range', 'message': 'effective_until must be >= effective_from'})

            is_active = _parse_bool(row.get('is_active'), default=True)

            if service_code and eff_from:
                key = (service_code, eff_from.isoformat())
                if key in seen_keys:
                    summary['duplicates_in_file'] += 1
                    row_errors.append({'field': 'service_code', 'code': 'duplicate', 'message': 'duplicate service_code + effective_from in file'})
                else:
                    seen_keys.add(key)

            if row_errors:
                summary['errors'] += len(row_errors)
                for e in row_errors:
                    issues.append({'row': idx, 'severity': 'error', **e})
                normalized.append({
                    'row': idx,
                    'service_code': service_code,
                    'external_code': external_code,
                    'effective_from': eff_from.isoformat() if eff_from else None,
                    'effective_until': eff_until.isoformat() if eff_until else None,
                    'is_active': bool(is_active),
                    'service_name': str(row.get('service_name') or '').strip() or None,
                    'category_name': str(row.get('category_name') or '').strip() or None,
                    'base_price': str(row.get('base_price') or '').strip() or None,
                })
                continue

            service_codes.add(service_code)
            normalized.append({
                'row': idx,
                'service_code': service_code,
                'external_code': external_code,
                'effective_from': eff_from.isoformat(),
                'effective_until': eff_until.isoformat() if eff_until else None,
                'is_active': bool(is_active),
                'service_name': str(row.get('service_name') or '').strip() or None,
                'category_name': str(row.get('category_name') or '').strip() or None,
                'base_price': str(row.get('base_price') or '').strip() or None,
            })

        facility = job.facility
        payer = job.payer

        services_by_code = {
            s.code: s
            for s in Service.objects.filter(facility=facility, code__in=service_codes).only('id', 'code')
        }
        missing_codes = [c for c in service_codes if c not in services_by_code]
        if missing_codes and not job.seed_services:
            for c in sorted(missing_codes)[:200]:
                summary['errors'] += 1
                issues.append({
                    'row': None,
                    'severity': 'error',
                    'field': 'service_code',
                    'code': 'not_found',
                    'message': f"service_code not found in facility: {c}",
                })
        if missing_codes and job.seed_services:
            summary['would_create_services'] = len(missing_codes)
            # Categories are optional; if provided, we can hint at category creates.
            cat_names = set()
            for r in normalized:
                if r.get('service_code') in missing_codes:
                    if r.get('category_name'):
                        cat_names.add(r['category_name'])
            if cat_names:
                existing_cats = set(
                    ServiceCategory.objects.filter(facility=facility, name__in=cat_names).values_list('name', flat=True)
                )
                summary['would_create_categories'] = len([n for n in cat_names if n not in existing_cats])

        # Determine would_create vs would_update for mappings (only for rows that have a known service).
        service_ids = [s.id for s in services_by_code.values()]
        existing = {}
        if service_ids:
            qs = PayerServiceCode.objects.filter(
                facility=facility,
                payer=payer,
                service_id__in=service_ids,
            ).only('id', 'service_id', 'external_code', 'effective_from', 'effective_until', 'is_active')
            for psc in qs:
                key = (psc.service_id, psc.effective_from.isoformat())
                existing.setdefault(key, []).append(psc)

        for r in normalized:
            if not r.get('effective_from') or not r.get('service_code') or not r.get('external_code'):
                continue
            svc = services_by_code.get(r['service_code'])
            if not svc:
                continue
            key = (svc.id, r['effective_from'])
            bucket = existing.get(key) or []
            if len(bucket) > 1:
                summary['errors'] += 1
                issues.append({
                    'row': r.get('row'),
                    'severity': 'error',
                    'field': 'service_code',
                    'code': 'duplicate_existing',
                    'message': 'multiple existing mappings found for service + effective_from; resolve manually',
                })
                continue
            if len(bucket) == 1:
                summary['would_update_mappings'] += 1
            else:
                summary['would_create_mappings'] += 1
            summary['rows_valid'] += 1

        parsed_bytes = json.dumps(normalized).encode('utf-8')

        PayerServiceCodeImportJob.objects.filter(id=job.id).update(
            status='preview_ready',
            parsed_payload_encrypted=encrypt_payload(parsed_bytes),
            summary=summary,
            issues=issues[:200],
            error_message='',
            processed_at=now,
            updated_at=now,
        )
    except Exception as e:
        PayerServiceCodeImportJob.objects.filter(id=job.id).update(
            status='failed',
            error_message=str(e)[:500],
            processed_at=now,
            updated_at=now,
        )
        raise


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def apply_payer_service_code_import_job(self, import_job_id: str) -> None:
    """
    Async apply for payer service code imports.

    Creates/updates PayerServiceCode mappings and optionally seeds Services.
    """
    import json

    job = PayerServiceCodeImportJob.objects.select_related('facility', 'payer', 'created_by').filter(id=import_job_id).first()
    if not job:
        return
    if job.status in ('applied', 'failed'):
        return
    if job.status not in ('preview_ready', 'applying'):
        return

    now = timezone.now()
    with transaction.atomic():
        locked = PayerServiceCodeImportJob.objects.select_for_update().get(id=job.id)
        if locked.status in ('applied', 'failed'):
            return
        if locked.status not in ('preview_ready', 'applying'):
            return
        if locked.status != 'applying':
            locked.status = 'applying'
            locked.save(update_fields=['status', 'updated_at'])

    try:
        parsed = decrypt_payload(job.parsed_payload_encrypted).decode('utf-8') if job.parsed_payload_encrypted else '[]'
        rows = json.loads(parsed) if parsed else []

        facility = job.facility
        payer = job.payer
        actor = job.created_by

        service_codes = sorted({(r.get('service_code') or '').strip() for r in rows if r.get('service_code')})
        existing_services = {
            s.code: s
            for s in Service.objects.filter(facility=facility, code__in=service_codes).select_related('category')
        }

        seeded_services = 0
        seeded_categories = 0

        if job.seed_services:
            default_category_name = "Imported (Unreviewed)"
            default_category, created = ServiceCategory.objects.get_or_create(
                facility=facility,
                name=default_category_name,
                defaults={
                    'description': 'Auto-created during payer service code import; review before activating.',
                    'is_active': True,
                    'created_by': actor,
                    'updated_by': actor,
                },
            )
            if created:
                seeded_categories += 1

            for code in service_codes:
                if code in existing_services:
                    continue
                # Best-effort details from import rows (first matching)
                row = next((r for r in rows if (r.get('service_code') or '').strip() == code), None) or {}
                category_name = (row.get('category_name') or '').strip() if isinstance(row.get('category_name'), str) else str(row.get('category_name') or '').strip()
                category = default_category
                if category_name:
                    category, cat_created = ServiceCategory.objects.get_or_create(
                        facility=facility,
                        name=category_name,
                        defaults={
                            'description': 'Auto-created during payer service code import; review before activating.',
                            'is_active': True,
                            'created_by': actor,
                            'updated_by': actor,
                        },
                    )
                    if cat_created:
                        seeded_categories += 1

                service_name = (row.get('service_name') or '').strip() if isinstance(row.get('service_name'), str) else str(row.get('service_name') or '').strip()
                if not service_name:
                    service_name = code
                base_price = Decimal('0.00')
                base_price_raw = row.get('base_price')
                if base_price_raw not in (None, ''):
                    try:
                        base_price = Decimal(str(base_price_raw))
                    except Exception:
                        base_price = Decimal('0.00')

                svc, created = Service.objects.get_or_create(
                    facility=facility,
                    code=code,
                    defaults={
                        'name': service_name[:100],
                        'description': 'Imported automatically; review pricing and activation status.',
                        'category': category,
                        'base_price': base_price,
                        'tax_rate': Decimal('0.00'),
                        'is_active': False,
                        'created_by': actor,
                        'updated_by': actor,
                    },
                )
                if created:
                    seeded_services += 1
                existing_services[code] = svc

        # Load existing mappings for payer/services.
        service_ids = [s.id for s in existing_services.values()]
        existing = {}
        if service_ids:
            qs = PayerServiceCode.objects.filter(
                facility=facility,
                payer=payer,
                service_id__in=service_ids,
            )
            for psc in qs:
                key = (psc.service_id, psc.effective_from.isoformat())
                existing.setdefault(key, []).append(psc)

        to_create = []
        to_update = []
        apply_issues = []

        for r in rows:
            code = (r.get('service_code') or '').strip()
            if not code:
                continue
            svc = existing_services.get(code)
            if not svc:
                apply_issues.append({
                    'row': r.get('row'),
                    'severity': 'error',
                    'field': 'service_code',
                    'code': 'not_found',
                    'message': f"service_code not found in facility: {code}",
                })
                continue

            eff_from = r.get('effective_from')
            if not eff_from:
                continue

            key = (svc.id, eff_from)
            bucket = existing.get(key) or []
            if len(bucket) > 1:
                apply_issues.append({
                    'row': r.get('row'),
                    'severity': 'error',
                    'field': 'service_code',
                    'code': 'duplicate_existing',
                    'message': 'multiple existing mappings found for service + effective_from; resolve manually',
                })
                continue

            ext = (r.get('external_code') or '').strip()
            eff_until = r.get('effective_until') or None
            is_active = bool(r.get('is_active', True))

            if len(bucket) == 1:
                obj = bucket[0]
                obj.external_code = ext
                obj.effective_until = _parse_date(eff_until) if eff_until else None
                obj.is_active = is_active
                obj.updated_by = actor
                obj.updated_at = now  # bulk_update won't auto_now
                to_update.append(obj)
            else:
                to_create.append(PayerServiceCode(
                    facility=facility,
                    payer=payer,
                    service=svc,
                    external_code=ext,
                    effective_from=_parse_date(eff_from),
                    effective_until=_parse_date(eff_until) if eff_until else None,
                    is_active=is_active,
                    created_by=actor,
                    updated_by=actor,
                ))

        if apply_issues:
            # Do not partially apply when apply-level errors exist.
            PayerServiceCodeImportJob.objects.filter(id=job.id).update(
                status='failed',
                error_message='apply_failed: fix errors and re-import',
                issues=(job.issues or [])[:200] + apply_issues[:200],
                applied_at=now,
                updated_at=now,
            )
            return

        if to_create:
            PayerServiceCode.objects.bulk_create(to_create, batch_size=500)
        if to_update:
            PayerServiceCode.objects.bulk_update(
                to_update,
                fields=['external_code', 'effective_until', 'is_active', 'updated_by', 'updated_at'],
                batch_size=500,
            )

        summary = job.summary or {}
        summary.update({
            'applied_create_mappings': len(to_create),
            'applied_update_mappings': len(to_update),
            'seeded_services': seeded_services,
            'seeded_categories': seeded_categories,
        })

        PayerServiceCodeImportJob.objects.filter(id=job.id).update(
            status='applied',
            summary=summary,
            applied_at=now,
            updated_at=now,
        )
    except Exception as e:
        PayerServiceCodeImportJob.objects.filter(id=job.id).update(
            status='failed',
            error_message=str(e)[:500],
            applied_at=now,
            updated_at=now,
        )
        raise


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def generate_nhis_claim_export(self, export_job_id: str) -> None:
    """
    Generate a Claim-it export payload (ZIP containing CSV files) and store it encrypted.
    """
    import csv
    import io
    import zipfile
    import hashlib

    job = NHISClaimExportJob.objects.select_related('batch', 'facility').filter(id=export_job_id).first()
    if not job:
        return
    if job.status not in ('pending',):
        return

    now = timezone.now()
    with transaction.atomic():
        locked = NHISClaimExportJob.objects.select_for_update().get(id=job.id)
        if locked.status != 'pending':
            return
        locked.status = 'running'
        locked.save(update_fields=['status', 'updated_at'])

    try:
        batch = NHISClaimBatch.objects.select_related('facility').get(id=job.batch_id)

        claims = list(
            Claim.objects.select_related(
                'invoice__patient',
                'invoice__patient_insurance__plan__provider',
                'invoice__patient_insurance__plan',
            ).prefetch_related(
                'invoice__items__service',
            ).filter(batch=batch)
        )

        payer_ids = set()
        service_ids = set()
        for claim in claims:
            inv = claim.invoice
            provider = inv.patient_insurance.plan.provider if inv.patient_insurance and inv.patient_insurance.plan else None
            if provider:
                payer_ids.add(provider.id)
            for item in inv.items.all():
                if item.service_id:
                    service_ids.add(item.service_id)

        codes = PayerServiceCode.objects.filter(
            facility=batch.facility,
            payer_id__in=list(payer_ids) if payer_ids else [],
            service_id__in=list(service_ids) if service_ids else [],
            is_active=True,
        )
        code_map = {}
        for c in codes:
            code_map.setdefault((c.payer_id, c.service_id), []).append(c)

        def _resolve_external_code(payer_id, service_id, inv_date):
            candidates = code_map.get((payer_id, service_id)) or []
            best = None
            for c in candidates:
                if c.effective_from and c.effective_from > inv_date:
                    continue
                if c.effective_until and c.effective_until < inv_date:
                    continue
                if not best or c.effective_from > best.effective_from:
                    best = c
            return best.external_code if best else ''

        claims_buf = io.StringIO()
        claims_writer = csv.writer(claims_buf)
        claims_writer.writerow([
            'claim_number',
            'invoice_number',
            'invoice_date',
            'patient_nhis_id',
            'payer_code',
            'plan_code',
            'claimed_amount',
            'approved_amount',
        ])

        items_buf = io.StringIO()
        items_writer = csv.writer(items_buf)
        items_writer.writerow([
            'claim_number',
            'invoice_number',
            'service_code',
            'service_name',
            'payer_service_code',
            'quantity',
            'unit_price',
            'line_total',
        ])

        for claim in claims:
            inv = claim.invoice
            patient = inv.patient
            plan = inv.patient_insurance.plan if inv.patient_insurance else None
            provider = plan.provider if plan else None

            claims_writer.writerow([
                claim.claim_number,
                inv.invoice_number,
                inv.invoice_date.isoformat() if inv.invoice_date else '',
                getattr(patient, 'nhis_id', '') or '',
                getattr(provider, 'code', '') or '',
                getattr(plan, 'code', '') or '',
                str(inv.insurance_amount),
                str(claim.approved_amount),
            ])

            payer_id = getattr(provider, 'id', None)
            for item in inv.items.all():
                ext_code = _resolve_external_code(payer_id, item.service_id, inv.invoice_date) if payer_id else ''
                svc = item.service
                items_writer.writerow([
                    claim.claim_number,
                    inv.invoice_number,
                    getattr(svc, 'code', '') or '',
                    getattr(svc, 'name', '') or '',
                    ext_code,
                    str(item.quantity),
                    str(item.unit_price),
                    str(item.total_price),
                ])

        # Package into a zip payload.
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('claims.csv', claims_buf.getvalue().encode('utf-8'))
            zf.writestr('items.csv', items_buf.getvalue().encode('utf-8'))

        payload_bytes = zip_buf.getvalue()
        checksum = hashlib.sha256(payload_bytes).hexdigest()

        NHISClaimExportJob.objects.filter(id=job.id).update(
            status='ready',
            payload_encrypted=encrypt_payload(payload_bytes),
            payload_checksum=checksum,
            error_message='',
            updated_at=now,
        )
    except Exception as e:
        NHISClaimExportJob.objects.filter(id=job.id).update(
            status='failed',
            error_message=str(e)[:500],
            updated_at=now,
        )
        raise


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def process_remittance_import_job(self, remittance_job_id: str) -> None:
    """
    Parse remittance CSV/XLSX and post insurance payments.
    """
    import csv
    import io
    import re
    from datetime import date, datetime
    from django.utils.dateparse import parse_date

    job = RemittanceImportJob.objects.select_related('facility', 'payer').filter(id=remittance_job_id).first()
    if not job:
        return
    if job.status not in ('pending',):
        return

    now = timezone.now()
    with transaction.atomic():
        locked = RemittanceImportJob.objects.select_for_update().get(id=job.id)
        if locked.status != 'pending':
            return
        locked.status = 'running'
        locked.save(update_fields=['status', 'updated_at'])

    try:
        raw = decrypt_payload(job.payload_encrypted) if job.payload_encrypted else b""
        file_name = (job.file_name or '').lower()

        def _norm_header(val) -> str:
            s = str(val or '').strip().lower()
            s = re.sub(r'[\s\-]+', '_', s)
            return s

        def _iter_source_rows():
            # CSV (default)
            if not file_name.endswith('.xlsx'):
                text = raw.decode('utf-8', errors='replace')
                reader = csv.DictReader(io.StringIO(text))
                for r in reader:
                    yield r
                return

            # XLSX (requires openpyxl)
            try:
                from openpyxl import load_workbook
            except ImportError as e:
                raise RuntimeError(
                    "XLSX remittance imports require openpyxl. Install openpyxl or convert the file to CSV."
                ) from e

            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header_row = next(it, None)
            if not header_row:
                return

            headers = [_norm_header(h) for h in header_row]
            for values in it:
                # Values-only rows can be shorter than headers.
                row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
                yield row

        rows = []
        claim_numbers = set()
        invoice_numbers = set()
        for row in _iter_source_rows():
            norm = {str(k or '').strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in (row or {}).items()}

            claim_number = norm.get('claim_number') or norm.get('claimno') or norm.get('claim')
            invoice_number = norm.get('invoice_number') or norm.get('invoice_no') or norm.get('invoice')
            paid_amount_raw = norm.get('paid_amount') or norm.get('amount_paid') or norm.get('amount') or norm.get('paid')
            paid_date_raw = norm.get('paid_date') or norm.get('payment_date') or norm.get('date')
            status_raw_val = norm.get('status') or norm.get('decision') or ''
            status_raw = (status_raw_val.strip().lower() if isinstance(status_raw_val, str) else str(status_raw_val).strip().lower())

            def _d(val):
                try:
                    return Decimal(str(val)) if val not in (None, '') else None
                except Exception:
                    return None

            paid_date = None
            if paid_date_raw not in (None, ''):
                if isinstance(paid_date_raw, datetime):
                    paid_date = paid_date_raw.date()
                elif isinstance(paid_date_raw, date):
                    paid_date = paid_date_raw
                else:
                    paid_date = parse_date(str(paid_date_raw))

            rows.append({
                'claim_number': str(claim_number).strip() if claim_number else '',
                'invoice_number': str(invoice_number).strip() if invoice_number else '',
                'paid_amount': _d(paid_amount_raw),
                'paid_date': paid_date,
                'status_raw': status_raw,
            })

            if claim_number:
                claim_numbers.add(str(claim_number).strip())
            if invoice_number:
                invoice_numbers.add(str(invoice_number).strip())

        claims = Claim.objects.select_related('invoice').filter(
            invoice__facility=job.facility,
            claim_number__in=list(claim_numbers) if claim_numbers else [],
        )
        claim_map = {c.claim_number: c for c in claims}

        invoices = Invoice.objects.filter(
            facility=job.facility,
            invoice_number__in=list(invoice_numbers) if invoice_numbers else [],
        )
        invoice_map = {i.invoice_number: i for i in invoices}

        for idx, row in enumerate(rows):
            claim = claim_map.get(row['claim_number']) if row['claim_number'] else None
            invoice = claim.invoice if claim else (invoice_map.get(row['invoice_number']) if row['invoice_number'] else None)

            match_status = 'unmatched'
            error_message = ''
            if row['paid_amount'] is None:
                match_status = 'error'
                error_message = 'paid_amount is required'
            elif not claim and not invoice:
                match_status = 'unmatched'
            else:
                # Determine denial/underpayment heuristics (no PHI).
                if row['paid_amount'] == 0 and row['status_raw'] in {'denied', 'rejected'}:
                    match_status = 'denied'
                else:
                    expected = None
                    try:
                        expected = (claim.approved_amount if claim and claim.approved_amount else None) or (
                            invoice.insurance_amount if invoice else None
                        )
                    except Exception:
                        expected = invoice.insurance_amount if invoice else None
                    if expected is not None and row['paid_amount'] < expected and row['paid_amount'] > 0:
                        match_status = 'underpaid'
                    else:
                        match_status = 'matched'

            line = RemittanceLine.objects.create(
                job=job,
                claim_number=row['claim_number'],
                invoice_number=row['invoice_number'],
                paid_amount=row['paid_amount'],
                paid_date=row['paid_date'],
                match_status=match_status,
                matched_claim=claim,
                matched_invoice=invoice,
                error_message=error_message,
            )

            if invoice and row['paid_amount'] and row['paid_amount'] > 0 and match_status in {'matched', 'underpaid'}:
                actor = job.created_by or getattr(invoice, 'updated_by', None) or getattr(invoice, 'created_by', None)

                reference_number = f"NHIS-{str(job.id).split('-')[0].upper()}-{idx}"
                payment = Payment(
                    invoice=invoice,
                    payment_date=(row['paid_date'] or now.date()),
                    amount=row['paid_amount'],
                    payer='insurance',
                    status='posted',
                    payment_method='insurance',
                    reference_number=reference_number,
                    notes=f"Remittance import {job.id}",
                    created_by=actor,
                    updated_by=actor,
                )
                payment.full_clean()
                payment.save()

                InsurancePosting.objects.create(
                    remittance_line=line,
                    payment=payment,
                    posted_at=now,
                    posted_by=actor,
                )

                _recompute_and_persist_invoice_status(invoice)

                if claim:
                    insurance_paid = Payment.objects.filter(
                        invoice=invoice, status='posted', payer='insurance'
                    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                    claim.paid_amount_total = insurance_paid
                    claim.last_paid_at = now
                    if (Decimal(str(invoice.insurance_amount)) - insurance_paid) <= 0:
                        claim.status = 'paid'
                    claim.save(update_fields=['paid_amount_total', 'last_paid_at', 'status', 'updated_at'])

        RemittanceImportJob.objects.filter(id=job.id).update(
            status='ready',
            processed_at=now,
            updated_at=now,
        )
    except Exception as e:
        RemittanceImportJob.objects.filter(id=job.id).update(
            status='failed',
            error_message=str(e)[:500],
            processed_at=now,
            updated_at=now,
        )
        raise


# =============================================================================
# Draft Invoice Auto-Sync (clinical -> financial)
# =============================================================================


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def sync_draft_invoice_for_encounter(self, encounter_id: str) -> None:
    """
    Sync the draft invoice for an encounter (if enabled/exists), based on current clinical state.

    PERF:
    - Run async; avoid blocking clinical endpoints.
    SECURITY:
    - Never log PHI; log opaque IDs only.
    """
    from apps.encounters.models import Encounter
    from apps.billing.services import DraftInvoiceSyncService

    encounter = Encounter.objects.select_related('facility', 'patient').filter(id=encounter_id).first()
    if not encounter:
        return

    service = DraftInvoiceSyncService()
    service.ensure_and_sync_for_encounter(
        encounter=encounter,
        actor=getattr(encounter, 'updated_by', None) or getattr(encounter, 'created_by', None),
    )


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def sync_draft_invoice_for_admission(self, admission_id: str) -> None:
    """
    Sync the draft invoice for an admission (if enabled/exists), based on current ward stay state.
    """
    from apps.wards.models import Admission
    from apps.billing.services import DraftInvoiceSyncService

    admission = Admission.objects.select_related('facility', 'patient').filter(id=admission_id).first()
    if not admission:
        return

    service = DraftInvoiceSyncService()
    service.ensure_and_sync_for_admission(
        admission=admission,
        actor=getattr(admission, 'updated_by', None) or getattr(admission, 'created_by', None),
    )


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def finalize_draft_invoice_for_encounter(self, encounter_id: str) -> None:
    """
    Finalize an encounter-linked draft invoice into a normal invoice (ready for payment/claims).
    """
    from apps.encounters.models import Encounter
    from apps.billing.services import DraftInvoiceSyncService

    encounter = Encounter.objects.select_related('facility', 'patient').filter(id=encounter_id).first()
    if not encounter:
        return

    service = DraftInvoiceSyncService()
    invoice = service.ensure_and_sync_for_encounter(
        encounter=encounter,
        actor=getattr(encounter, 'updated_by', None) or getattr(encounter, 'created_by', None),
    )
    service.finalize_invoice(
        invoice=invoice,
        actor=getattr(encounter, 'updated_by', None) or getattr(encounter, 'created_by', None),
    )


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def finalize_draft_invoice_for_admission(self, admission_id: str) -> None:
    """
    Finalize an admission-linked draft invoice when the patient is discharged.
    """
    from apps.wards.models import Admission
    from apps.billing.services import DraftInvoiceSyncService

    admission = Admission.objects.select_related('facility', 'patient').filter(id=admission_id).first()
    if not admission:
        return

    service = DraftInvoiceSyncService()
    invoice = service.ensure_and_sync_for_admission(
        admission=admission,
        actor=getattr(admission, 'updated_by', None) or getattr(admission, 'created_by', None),
    )
    service.finalize_invoice(
        invoice=invoice,
        actor=getattr(admission, 'updated_by', None) or getattr(admission, 'created_by', None),
    )
